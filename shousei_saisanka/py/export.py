from imp import json, linked, ConvertLi
from CRUD import MRS, N2_SAISANKA, QS_FURNACE, CHECKBOXES
import openpyxl as px # https://community.esri.com/t5/python-questions/how-to-update-existing-excel-xlsx-file-using/td-p/1065848
from shutil import copy2
from duplicate_sheet import copy_sheet_ws

tableName, data = linked
# data = "[O2#3,X,2,2022,false]"
# tableName = MRS
# tableName = QS_FURNACE
# tableName = N2_SAISANKA
# tableName = "o2_saisanka"
# tableName = "o2_fast_cooling"
# tableName = "o2_low_temperature"
# tableName = "qh_furnace"

searchLi = ConvertLi(data)
furnace_no, day, month, year, fileNameOnly = searchLi
niceDate = "" if day == "X" else f"{day}-"
niceDate += f"{month}-{year}"
# QUERY_DATE, QUERY_MAC = status.split(",")
# QUERY_MAC = int(QUERY_MAC)

# QUERY_DATE = '2022-01-10'
# QUERY_MAC = 5
tnFn = {
    "mrs" : "MRS Shousei",
    "qs_furnace": "QSS Furnace",
    "n2_saisanka": "N2 Saisanka",
    "o2_saisanka": "O2 Normal",
    "o2_fast_cooling": "O2 Fast Cooling",
    "o2_low_temperature": "O2 Low-low Temperature",
    "qh_furnace": "QH Furnace",
}

ORI_FILE = f"./excel/{tableName}.xlsx"
NEW_FILENAME = f"{tnFn[tableName]} ({furnace_no}) ({niceDate})".replace("#", "_")
NEW_PATH = "./export/"
NEW_FILE = f"{NEW_PATH}{NEW_FILENAME}.xlsx"

if fileNameOnly == "true":
    print(NEW_FILENAME)
    exit()

if tableName == MRS:
    FUR_NO_ROW = 2
    FUR_NO_COL = 11
    MONTH_ROW = MONTH_COL = YEAR_ROW = YEAR_COL = 0
    ONE_PAGE_NUM = 20
    START_ROW = 4
    START_COL = 3
    FIELDS = ("date", "start_dt", "end_dt", "item_name", "key_no", "lot_no", "lot_size", 
    "quantity_plan", "quantity_actual", "pattern_no", "quantity_stick", "rate_stick", 
    "narabe_pic", "collect_pic", "grp")
elif tableName == QS_FURNACE:
    FUR_NO_ROW = 5
    FUR_NO_COL = 2
    MONTH_ROW = 7
    MONTH_COL = 2
    YEAR_ROW = 7
    YEAR_COL = 4
    ONE_PAGE_NUM = 25
    START_ROW = 15
    START_COL = 2
    FIELDS = ("inspection_date", "inspection_time", "item_type", "lot_no", "lot_quantity", "item", 
    "slit", "slit_nfmd", "dash", "temp", "tube_nfmd", "feeder_nfmd", "tray_nfmd", 
    "done_by", "checked_by", "verified_by", "remark")
else:
    FUR_NO_ROW = 3
    FUR_NO_COL = 1
    MONTH_ROW = MONTH_COL = 0
    YEAR_ROW = 3
    YEAR_COL = 3
    ONE_PAGE_NUM = 25
    START_ROW = 9
    START_COL = 1
    FIELDS = ["ticket_no", "item_name", "lot_no", "lot_size", "setter_pilling", "chip_layer", "input_date", "input_time", 
    "output_date", "output_time", "narabe_quantity", "narabe_pic", "collect_pic", "confirmation"]
    if tableName == N2_SAISANKA:
        for remField in ("setter_pilling", "confirmation"): 
            FIELDS.remove(remField)
        FIELDS.append("o2_free")
        
def LoadWBWS():
    wb = px.load_workbook(NEW_FILE)
    ws = wb.worksheets[0]
    return wb, ws
def SaveCloseWB(wb): 
    wb.save(NEW_FILE)
    wb.close()

copy2(ORI_FILE, NEW_FILE)
wb, ws = LoadWBWS()
ws.cell(row=FUR_NO_ROW, column=FUR_NO_COL).value += furnace_no
if MONTH_ROW or MONTH_COL:
    ws.cell(row=MONTH_ROW, column=MONTH_COL).value += month
if YEAR_ROW or YEAR_COL:
    ws.cell(row=YEAR_ROW, column=YEAR_COL).value += year
SaveCloseWB(wb)

with open(f"{NEW_PATH}{NEW_FILENAME}.json") as f:
    docs = json.load(f)
    
numDocs = len(docs)
numSheet = int("{:.2f}".format(numDocs / ONE_PAGE_NUM).split('.')[0])
if numDocs % ONE_PAGE_NUM != 0: numSheet += 1
for i in range(1, numSheet): 
    # copy_sheet_ws(NEW_FILE, ws.title, f"{ws.title} ({i+1})")
    # invoke copy_worksheet method to clone source sheet.
    wb, ws = LoadWBWS()
    target_sheet = wb.copy_worksheet(ws)
    target_sheet.title = f"{ws.title} ({i+1})"
    target_sheet.print_area = ws.print_area
    target_sheet.sheet_view.view = ws.sheet_view.view
    target_sheet.sheet_view.showGridLines = False
    SaveCloseWB(wb)

wb, ws = LoadWBWS()
index = 0
row = START_ROW
docs.reverse()
for d in docs:    
    keyCol = {}
    c = START_COL
    for f in FIELDS:
        keyCol[f] = c
        c += 1

    if tableName == MRS:
        stick = d["quantity_stick"]
        d["rate_stick"] = stick / d["lot_size"] * 100 if stick else 0
    elif tableName == QS_FURNACE:
        if d["item"] == "063": 
            keyCol["temp"] -= 1
            keyCol["dash"] += 1
        d["dash"] = "-"
        d["temp"] = str(d["temp"])
        d["temp"] += " °C"
    elif tableName == N2_SAISANKA:
        d["o2_free"] = "Yes" if d["o2_free"] else "No"

    for c in CHECKBOXES:
        if c in d and "nfmd" in c:
            d[c] = "O" if d[c] else "X"
    cfm = "confirmation"
    if cfm in d: d[cfm] = "/" if d[cfm] else ""
    
    for key in keyCol: 
        ws.cell(row=row, column=keyCol[key]).value = d[key]
    row += 1
    index += 1
    if index != 0 and index % ONE_PAGE_NUM == 0: 
        row = START_ROW
        # Next worksheet
        try: ws = wb.worksheets[wb.worksheets.index(ws) + 1]
        except: pass
SaveCloseWB(wb)