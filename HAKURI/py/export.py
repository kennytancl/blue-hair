from imp import linked, Counter, datetime
from typing import cast
from CRUD import read as CRUDread
from CRUD import FMT_DATE
import openpyxl as px # https://community.esri.com/t5/python-questions/how-to-update-existing-excel-xlsx-file-using/td-p/1065848
from shutil import copy2
from duplicate_sheet import copy_sheet_ws

status, data = linked # linked = [status, data]
QUERY_DATE, QUERY_MAC = status.split(",")
QUERY_MAC = int(QUERY_MAC)

# QUERY_DATE = '2022-01-10'
# QUERY_MAC = 5

ONE_PAGE_NUM = 45
ORI_FILE = './excel/record.xlsx'
NEW_FILE = './excel/HAKURI_RECORD({})(Mac {}).xlsx'\
    .format(QUERY_DATE, QUERY_MAC)
def SHEET_NAME(num): return 'FORM CEE-H10306R2-{}'.format(num)

copy2(ORI_FILE, NEW_FILE)

def LoadWBWS():
    wb = px.load_workbook(r"{}".format(NEW_FILE))
    ws = wb[SHEET_NAME(1)]
    return wb, ws

def SaveWB(wb):
    wb.save(r"{}".format(NEW_FILE))

class Item:    
    def __init__(self, itemName = "", lotNo = ""):
        self.itemName = itemName
        self.lotNo = lotNo
        
    def __eq__(self, other):
        return self.itemName == other.itemName and self.lotNo == other.lotNo

wb, ws = LoadWBWS()
ws.cell(row=4, column=2).value = "Date: {}".format(datetime.strptime(QUERY_DATE, FMT_DATE).strftime('%d.%m.%Y'))
ws.cell(row=5, column=2).value = "Machine No: {}".format(QUERY_MAC)
SaveWB(wb)

batch = CRUDread("READ_start_dt", "[{}]".format(QUERY_DATE))
count = 0
for b in batch: 
    if b["machine_no"] != QUERY_MAC: continue
    prevItem = Item()
    for i in b["items"]:
        currItem = Item(i["item_name"], i["lot_no"])
        if prevItem == currItem: continue
        count += 1
        prevItem = currItem

numSheet = int("{:.2f}".format(count / ONE_PAGE_NUM).split('.')[0])
if count % ONE_PAGE_NUM != 0: numSheet += 1
for c in range(1, numSheet):
    copy_sheet_ws(NEW_FILE, SHEET_NAME(1), SHEET_NAME(c + 1))

wb, ws = LoadWBWS()
sheet_num = 1
row = 9
index = 0
batch.reverse()
for b in batch: # Write method, row, column, value
    if b["machine_no"] != QUERY_MAC: continue
    names = []
    for i in b["items"]: names.append(i["item_name"])
    countItems = Counter(names)
    prevItem = Item()
    for i in b["items"]:
        currItem = Item(i["item_name"], i["lot_no"])
        if prevItem == currItem: continue
        
        countItemNames = countItems[i["item_name"]]
        similarItem = countItemNames != 1
        ws.cell(row=row, column=2).value = i["item_name"]
        ws.cell(row=row, column=3).value = i["lot_no"]

        qty = 0
        for temp in b["items"]:
            tempItem = Item(temp["item_name"], temp["lot_no"])
            if tempItem == currItem: qty += 1
        ws.cell(row=row, column=4).value = qty

        tank_no = [str(i["tank_no"])]
        for t in range(qty - 1):
            tank_no.append(str(i["tank_no"] + t + 1))
        ws.cell(row=row, column=5).value = ', '.join(tank_no)
        ws.cell(row=row, column=6).value = b["start_dt"].split('<br>')[1]
        finish_time = b["finish_dt"].split('<br>')[1] if b["finish_dt"] else "None"
        ws.cell(row=row, column=7).value = finish_time
        b["end_pic"] = b["end_pic"] if b["end_pic"] else "None"
        ws.cell(row=row, column=8).value = "{}, {}".format(b["start_pic"], b["end_pic"])
        row += 1
        index += 1

        if index != 0 and index % ONE_PAGE_NUM == 0: 
            row = 9
            sheet_num += 1
            try: ws = wb[SHEET_NAME(sheet_num)]
            except: pass

        prevItem = currItem
        
SaveWB(wb)
