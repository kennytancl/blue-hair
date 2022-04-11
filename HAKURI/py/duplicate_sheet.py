# https://www.dev2qa.com/how-to-use-python-openpyxl-to-copy-excel-sheet-data-in-same-and-different-excel-file/

import openpyxl, os
from openpyxl import Workbook
from openpyxl import load_workbook

def if_file_exist(file_path):
    return os.path.exists(file_path)
    
def if_excel_sheet_exist(work_book, work_sheet_name):
    sheet_names_list = work_book.sheetnames
    for sheet_name in sheet_names_list:
        return sheet_name == work_sheet_name

def copy_sheet(file_path, source_sheet_name, target_sheet_name):
    if if_file_exist(file_path):
        # print("File ", file_path, " exist.")
        # load the excel file return Workbook object
        work_book = load_workbook(file_path)
        if if_excel_sheet_exist(work_book, source_sheet_name):
            # print("Source excel sheet ", source_sheet_name, " exist.")  
            # get source Worksheet object.
            source_work_sheet = work_book[source_sheet_name]
            # initialise the target_work_sheet to None.
            target_work_sheet = None
            # if target excel sheet exist in the excel file.
            if if_excel_sheet_exist(work_book, target_sheet_name):
                # assign the target Worksheet.
                target_work_sheet = work_book[target_sheet_name]
            else:
                # create a new Worksheet object.
                target_work_sheet = work_book.create_sheet(target_sheet_name)
            # loop the source excel sheet rows.
            row_number = 1
            for row in source_work_sheet.iter_rows():
                # loop the cell in the row.
                cell_column_number = 1
                for cell in row:
                    # create a cell in target work sheet.
                    target_cell = target_work_sheet.cell(row = row_number, column = cell_column_number, value = cell.value)
                    cell_column_number += 1
                row_number += 1
            # save the excel file.
            work_book.save(file_path)

def copy_sheet_ws(file_path, source_sheet_name, target_sheet_name):
    if if_file_exist(file_path):
        # load the excel file return Workbook object
        work_book = load_workbook(file_path)
        if if_excel_sheet_exist(work_book, source_sheet_name):
            source_sheet = work_book[source_sheet_name]
            # invoke copy_worksheet method to clone source sheet.
            target_sheet = work_book.copy_worksheet(source_sheet)
            target_sheet.title = target_sheet_name    
            # save the excel file.
            work_book.save(file_path)

source_file_path = './excel/record.xlsx'
source_sheet_name = 'FORM CEE-H10306R2'
target_sheet_name = 'User Account New'
# copy_sheet_ws(source_file_path, source_sheet_name, target_sheet_name)
copy_sheet(source_file_path, source_sheet_name, target_sheet_name)